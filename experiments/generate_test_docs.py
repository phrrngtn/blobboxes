"""Generate synthetic test documents with known table structures.

Each document embeds tables with known ground truth (column positions,
header rows, data types). The generator returns the ground truth alongside
the files so we can validate geometry-based column detection.
"""

import json
from pathlib import Path
from dataclasses import dataclass, field, asdict

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

import openpyxl
from openpyxl.styles import Font as XlFont, PatternFill, Alignment

OUT = Path("/Users/paulharrington/checkouts/blobboxes/test_data/synthetic")
OUT.mkdir(parents=True, exist_ok=True)


@dataclass
class GroundTruthColumn:
    name: str
    col_index: int  # 0-based
    dtype: str      # "string", "numeric", "date", "currency"


@dataclass
class GroundTruthTable:
    table_id: str
    description: str
    header_row: int     # 0-based row index within the table
    num_data_rows: int
    num_cols: int
    columns: list       # list of GroundTruthColumn dicts
    difficulty: str     # "easy", "medium", "hard"


# ── Table definitions ────────────────────────────────────────────────

SALES_TABLE = {
    "table_id": "sales_simple",
    "description": "Simple sales table with clear headers and uniform data",
    "difficulty": "easy",
    "headers": ["Region", "Product", "Q1 Sales", "Q2 Sales", "Q3 Sales", "Total"],
    "col_types": ["string", "string", "currency", "currency", "currency", "currency"],
    "rows": [
        ["North", "Widget A", "$12,400", "$13,200", "$14,800", "$40,400"],
        ["North", "Widget B", "$8,900", "$9,100", "$10,200", "$28,200"],
        ["South", "Widget A", "$15,600", "$14,900", "$16,100", "$46,600"],
        ["South", "Widget B", "$7,200", "$8,400", "$9,800", "$25,400"],
        ["East", "Widget A", "$11,300", "$12,700", "$13,500", "$37,500"],
        ["East", "Widget B", "$6,800", "$7,200", "$8,100", "$22,100"],
        ["West", "Widget A", "$14,100", "$15,300", "$16,900", "$46,300"],
        ["West", "Widget B", "$9,500", "$10,100", "$11,200", "$30,800"],
    ],
}

EMPLOYEE_TABLE = {
    "table_id": "employee_directory",
    "description": "Employee directory with mixed types",
    "difficulty": "easy",
    "headers": ["ID", "Name", "Department", "Start Date", "Salary"],
    "col_types": ["numeric", "string", "string", "date", "currency"],
    "rows": [
        ["1001", "Alice Johnson", "Engineering", "2019-03-15", "$95,000"],
        ["1002", "Bob Smith", "Marketing", "2020-07-01", "$78,000"],
        ["1003", "Carol Davis", "Engineering", "2018-11-20", "$102,000"],
        ["1004", "David Lee", "Sales", "2021-01-10", "$72,000"],
        ["1005", "Eve Martinez", "Engineering", "2022-05-22", "$88,000"],
        ["1006", "Frank Wilson", "Marketing", "2019-09-14", "$81,000"],
    ],
}

# Medium: table with a title row above the headers
INVENTORY_TABLE = {
    "table_id": "inventory_with_title",
    "description": "Table preceded by a title and subtitle — headers are NOT row 0",
    "difficulty": "medium",
    "title": "Warehouse Inventory Report",
    "subtitle": "As of March 2026 — All locations",
    "headers": ["SKU", "Description", "Location", "Qty On Hand", "Reorder Point", "Unit Cost"],
    "col_types": ["string", "string", "string", "numeric", "numeric", "currency"],
    "rows": [
        ["WH-001", "Hex Bolt M8x30", "Aisle 3, Bin 12", "4500", "1000", "$0.12"],
        ["WH-002", "Flat Washer M8", "Aisle 3, Bin 13", "12000", "3000", "$0.03"],
        ["WH-003", "Lock Nut M8", "Aisle 3, Bin 14", "3200", "800", "$0.08"],
        ["WH-004", "Spring Pin 4x30", "Aisle 5, Bin 1", "890", "500", "$0.45"],
        ["WH-005", "Cotter Pin 3x25", "Aisle 5, Bin 2", "2100", "600", "$0.15"],
    ],
}

# Hard: two tables on same page with different column counts
MULTI_TABLE = {
    "table_id": "multi_table_page",
    "description": "Two tables on same page — different column counts, separated by text",
    "difficulty": "hard",
    "table_a": {
        "sub_id": "summary_metrics",
        "title": "Summary Metrics",
        "headers": ["Metric", "Value"],
        "col_types": ["string", "numeric"],
        "rows": [
            ["Total Revenue", "$2,450,000"],
            ["Total Expenses", "$1,890,000"],
            ["Net Income", "$560,000"],
            ["Headcount", "142"],
            ["Revenue per Employee", "$17,254"],
        ],
    },
    "table_b": {
        "sub_id": "quarterly_breakdown",
        "title": "Quarterly Breakdown",
        "headers": ["Quarter", "Revenue", "Expenses", "Margin %"],
        "col_types": ["string", "currency", "currency", "numeric"],
        "rows": [
            ["Q1 2025", "$580,000", "$470,000", "19.0%"],
            ["Q2 2025", "$620,000", "$480,000", "22.6%"],
            ["Q3 2025", "$640,000", "$460,000", "28.1%"],
            ["Q4 2025", "$610,000", "$480,000", "21.3%"],
        ],
    },
}

# Hard: sparse table with merged-looking header and footnotes
SPARSE_TABLE = {
    "table_id": "sparse_with_notes",
    "description": "Table with blank cells, footnote markers, and irregular spacing",
    "difficulty": "hard",
    "headers": ["Country", "Capital", "Population (M)", "GDP ($B)", "HDI"],
    "col_types": ["string", "string", "numeric", "numeric", "numeric"],
    "rows": [
        ["United States", "Washington, D.C.", "331.9", "25,462", "0.921"],
        ["China", "Beijing", "1,412.0", "17,963", "0.768"],
        ["Japan", "Tokyo", "125.7", "4,231", "0.925"],
        ["Germany", "Berlin", "83.2", "4,072", "0.942"],
        ["India", "New Delhi", "1,408.0", "3,385", "0.633"],
        ["United Kingdom", "London", "67.3", "3,070", "0.929"],
        ["France", "Paris", "67.8", "2,778", "0.903"],
    ],
    "footnotes": [
        "* Population figures: UN World Population Prospects 2022",
        "** GDP: World Bank 2022 estimates, current USD",
        "*** HDI: UNDP Human Development Report 2021/22",
    ],
}


def build_ground_truth(table_def, header_row_offset=0):
    """Build GroundTruthTable from a table definition dict."""
    gt = GroundTruthTable(
        table_id=table_def["table_id"],
        description=table_def["description"],
        header_row=header_row_offset,
        num_data_rows=len(table_def["rows"]),
        num_cols=len(table_def["headers"]),
        columns=[
            asdict(GroundTruthColumn(name=h, col_index=i, dtype=t))
            for i, (h, t) in enumerate(zip(table_def["headers"], table_def["col_types"]))
        ],
        difficulty=table_def["difficulty"],
    )
    return gt


# ── PDF generation ───────────────────────────────────────────────────

def make_pdf_table_style(has_header=True):
    style_cmds = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    if has_header:
        style_cmds += [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]
    return TableStyle(style_cmds)


def generate_pdf_simple(table_def, filename):
    """Single clear table, easy case."""
    path = OUT / filename
    doc = SimpleDocTemplate(str(path), pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    elements = []

    data = [table_def["headers"]] + table_def["rows"]
    t = Table(data, repeatRows=1)
    t.setStyle(make_pdf_table_style())
    elements.append(t)

    doc.build(elements)
    return build_ground_truth(table_def, header_row_offset=0)


def generate_pdf_with_title(table_def, filename):
    """Table with title/subtitle above — header is NOT the first text."""
    path = OUT / filename
    doc = SimpleDocTemplate(str(path), pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(table_def["title"], styles["Title"]))
    elements.append(Paragraph(table_def["subtitle"], styles["Normal"]))
    elements.append(Spacer(1, 0.3*inch))

    data = [table_def["headers"]] + table_def["rows"]
    t = Table(data, repeatRows=1)
    t.setStyle(make_pdf_table_style())
    elements.append(t)

    doc.build(elements)
    # header_row_offset accounts for title + subtitle bboxes above the table
    return build_ground_truth(table_def, header_row_offset=0)


def generate_pdf_multi(multi_def, filename):
    """Two tables on one page separated by a paragraph."""
    path = OUT / filename
    doc = SimpleDocTemplate(str(path), pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    elements = []

    # Table A
    ta_def = multi_def["table_a"]
    elements.append(Paragraph(ta_def["title"], styles["Heading2"]))
    elements.append(Spacer(1, 0.1*inch))
    data_a = [ta_def["headers"]] + ta_def["rows"]
    t_a = Table(data_a, repeatRows=1)
    t_a.setStyle(make_pdf_table_style())
    elements.append(t_a)

    elements.append(Spacer(1, 0.4*inch))
    elements.append(Paragraph(
        "The quarterly breakdown shows consistent growth across all periods.",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 0.2*inch))

    # Table B
    tb_def = multi_def["table_b"]
    elements.append(Paragraph(tb_def["title"], styles["Heading2"]))
    elements.append(Spacer(1, 0.1*inch))
    data_b = [tb_def["headers"]] + tb_def["rows"]
    t_b = Table(data_b, repeatRows=1)
    t_b.setStyle(make_pdf_table_style())
    elements.append(t_b)

    doc.build(elements)

    gt_a = GroundTruthTable(
        table_id=ta_def["sub_id"], description=ta_def["title"],
        header_row=0, num_data_rows=len(ta_def["rows"]),
        num_cols=len(ta_def["headers"]),
        columns=[asdict(GroundTruthColumn(name=h, col_index=i, dtype=t))
                 for i, (h, t) in enumerate(zip(ta_def["headers"], ta_def["col_types"]))],
        difficulty="hard",
    )
    gt_b = GroundTruthTable(
        table_id=tb_def["sub_id"], description=tb_def["title"],
        header_row=0, num_data_rows=len(tb_def["rows"]),
        num_cols=len(tb_def["headers"]),
        columns=[asdict(GroundTruthColumn(name=h, col_index=i, dtype=t))
                 for i, (h, t) in enumerate(zip(tb_def["headers"], tb_def["col_types"]))],
        difficulty="hard",
    )
    return [gt_a, gt_b]


def generate_pdf_with_footnotes(table_def, filename):
    """Table with footnotes below — tests bottom boundary detection."""
    path = OUT / filename
    doc = SimpleDocTemplate(str(path), pagesize=letter,
                            leftMargin=0.75*inch, rightMargin=0.75*inch,
                            topMargin=0.75*inch, bottomMargin=0.75*inch)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Country Statistics", styles["Title"]))
    elements.append(Spacer(1, 0.2*inch))

    data = [table_def["headers"]] + table_def["rows"]
    t = Table(data, repeatRows=1)
    t.setStyle(make_pdf_table_style())
    elements.append(t)

    elements.append(Spacer(1, 0.15*inch))
    for fn in table_def["footnotes"]:
        elements.append(Paragraph(fn, styles["Normal"]))

    doc.build(elements)
    return build_ground_truth(table_def, header_row_offset=0)


# ── XLSX generation ──────────────────────────────────────────────────

def generate_xlsx_simple(table_def, filename):
    """Simple XLSX with header row + data."""
    path = OUT / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    header_font = XlFont(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    for c, h in enumerate(table_def["headers"], 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for r, row in enumerate(table_def["rows"], 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(path)
    return build_ground_truth(table_def, header_row_offset=0)


def generate_xlsx_with_title(table_def, filename):
    """XLSX with title in row 1, blank row, then headers in row 3."""
    path = OUT / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # Title in A1
    ws.cell(row=1, column=1, value=table_def["title"]).font = XlFont(bold=True, size=14)
    # Subtitle in A2
    ws.cell(row=2, column=1, value=table_def["subtitle"]).font = XlFont(italic=True, size=10)
    # Blank row 3
    # Headers in row 4
    header_font = XlFont(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for c, h in enumerate(table_def["headers"], 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for r, row in enumerate(table_def["rows"], 5):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(path)
    # header_row is 3 (0-based) because rows 0,1,2 are title/subtitle/blank
    gt = build_ground_truth(table_def, header_row_offset=3)
    return gt


def generate_xlsx_multi(multi_def, filename):
    """Two tables on same sheet, separated by blank rows."""
    path = OUT / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    header_font = XlFont(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    # Table A starting at row 1
    ta = multi_def["table_a"]
    ws.cell(row=1, column=1, value=ta["title"]).font = XlFont(bold=True, size=12)
    for c, h in enumerate(ta["headers"], 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r, row in enumerate(ta["rows"], 3):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    # Blank rows, then Table B
    start_b = 3 + len(ta["rows"]) + 2  # 2 blank rows
    tb = multi_def["table_b"]
    ws.cell(row=start_b, column=1, value=tb["title"]).font = XlFont(bold=True, size=12)
    for c, h in enumerate(tb["headers"], 1):
        cell = ws.cell(row=start_b + 1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for r, row in enumerate(tb["rows"], start_b + 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(path)
    return multi_def  # ground truth is the full definition


# ── Main ─────────────────────────────────────────────────────────────

def main():
    ground_truth = {}

    # PDF documents
    gt = generate_pdf_simple(SALES_TABLE, "01_sales_simple.pdf")
    ground_truth["01_sales_simple.pdf"] = [asdict(gt)]

    gt = generate_pdf_simple(EMPLOYEE_TABLE, "02_employee_simple.pdf")
    ground_truth["02_employee_simple.pdf"] = [asdict(gt)]

    gt = generate_pdf_with_title(INVENTORY_TABLE, "03_inventory_titled.pdf")
    ground_truth["03_inventory_titled.pdf"] = [asdict(gt)]

    gts = generate_pdf_multi(MULTI_TABLE, "04_multi_table.pdf")
    ground_truth["04_multi_table.pdf"] = [asdict(g) for g in gts]

    gt = generate_pdf_with_footnotes(SPARSE_TABLE, "05_countries_footnotes.pdf")
    ground_truth["05_countries_footnotes.pdf"] = [asdict(gt)]

    # XLSX documents
    gt = generate_xlsx_simple(SALES_TABLE, "06_sales_simple.xlsx")
    ground_truth["06_sales_simple.xlsx"] = [asdict(gt)]

    gt = generate_xlsx_with_title(INVENTORY_TABLE, "07_inventory_titled.xlsx")
    ground_truth["07_inventory_titled.xlsx"] = [asdict(gt)]

    # Save ground truth
    gt_path = OUT / "ground_truth.json"
    with open(gt_path, "w") as f:
        json.dump(ground_truth, f, indent=2)

    print(f"Generated {len(ground_truth)} test documents in {OUT}/")
    for name, tables in ground_truth.items():
        for t in tables:
            print(f"  {name}: {t['table_id']} ({t['difficulty']}) "
                  f"— {t['num_cols']} cols x {t['num_data_rows']} rows")


if __name__ == "__main__":
    main()
